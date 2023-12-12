import pytesseract
from pdf2image import convert_from_path
import cv2
import os
import shutil
import openpyxl


def crop_pdf(pdf_path, crop_area, output_path):
    # Convert the first page of the PDF to an image
    images = convert_from_path(pdf_path, first_page=1, last_page=1)

    if images:
        # Crop the image
        cropped_image = images[0].crop(crop_area)  # crop_area is (left, upper, right, lower)
        cropped_image.save(output_path)


def preprocess_and_ocr(image_path):
    # Read the image
    img = cv2.imread(image_path)

    # Convert to grayscale
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)

    # Apply Gaussian blur
    blur = cv2.GaussianBlur(gray, (5,5), 0)

    # Apply thresholding
    _, thresh = cv2.threshold(blur, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)

    # OCR with Tesseract
    text = pytesseract.image_to_string(thresh, lang='chi_sim')

    return text

def scan_for_pdfs(directory):
    pdf_files = []
    xlsx_files = []
    for root, dirs, files in os.walk(directory):
        for file in files:
            if file.endswith('.pdf'):
                pdf_files.append(os.path.join(root, file))
            if file.endswith('.xlsx'):
                xlsx_files.append(os.path.join(root, file))

    return pdf_files, xlsx_files

def write_to_excel(file_path, row=9, column=1, data=''):
    # Load or create an Excel workbook
    try:
        workbook = openpyxl.load_workbook(file_path)
    except FileNotFoundError:
        workbook = openpyxl.Workbook()

    # Select the active sheet or create a new one
    if workbook.sheetnames:
        sheet = workbook.active
    else:
        sheet = workbook.create_sheet("Sheet1")

    # Write data to the 9th row, 1st column
    sheet.cell(row=row, column=column, value=data)

    # Save the workbook
    workbook.save(file_path)


def add_header(xlsx_files):
    # add header
    for xlsx in xlsx_files:
        count = 1
        for key, value in extract_place.items():
            write_to_excel(xlsx, row=9, column=count, data=key)
            count += 1


def excute_extract(pdf_list, xlsx_files):
    add_header(xlsx_files)
    # extract text
    for i in range(len(pdf_list)):
        pdf = pdf_list[i]
        xlsx = xlsx_files[xlsx_files.index(pdf.replace('.pdf', '.xlsx'))]
        count = 1
        print('-------------------')
        print(xlsx)
        for key, value in extract_place.items():
            # Create a folder to store the cropped images
            folder_path = 'temp/'  # Replace with your desired path
            os.makedirs(folder_path, exist_ok=True)  # This will create the folder if it doesn't exist

            # 然后分别切割并且识别
            output_path = folder_path + key + '.png'
            crop_pdf(pdf, value, output_path=output_path)
            text = preprocess_and_ocr(output_path)
            text = text.replace('劳务+', '')
            text = text.replace('劳务*', '')
            print(text)

            # write text to excel
            write_to_excel(xlsx, row=10, column=count, data=text)
            count += 1

            # Delete the folder, 抠图的时候可以不删除，注释掉下面语句
            shutil.rmtree(folder_path)  # This will delete the folder and all its contents


if __name__ == '__main__':
    # 扫描目录获得pdf list
    pdf_list, xlsx_files = scan_for_pdfs('/Users/zhoujianyi/code/python/Spider-Demo1/extract_text/ssss')  # Replace with your directory path
    print(pdf_list)
    print(xlsx_files)

    # TODO: 自己添加需要提取的字段，(left, upper, right, lower)
    extract_place = {'名称': (150, 250, 700, 300), '项目名称': (35, 440, 350, 650)}

    excute_extract(pdf_list, xlsx_files)

